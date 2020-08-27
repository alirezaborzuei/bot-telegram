from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters,
                          ConversationHandler)
from openpyxl import load_workbook
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
updater = Updater("1080958406:AAEjfkoV019XFaDgwnA-9ri89FnBfGphwrE", use_context=True)
FileName = ""
A, B = range(2)

def make(update, filename):
    global FileName
    FileName = filename
    global sheet
    book = load_workbook('dataset.xlsx')
    sheets = book.sheetnames
    i = 0
    for i in range(0, len(sheets)):
        if sheets[i] == filename:
            sheet = book.get_sheet_by_name(filename)

    if sheets[i] != filename:
        update.message.reply_text('آیدی به این اسم وجود ندارد و آیدی درست را وارد نمائید:')




def id_channel(update, context):
    chat_id = update.effective_chat.id
    context.bot.send_message(chat_id, text="آیدی کانال یا گروه یا شخصی که میخواهید محصولات را برای آن ایجاد کنید بنویسید.")
    return A


def get_id_channel(update, context):
    make(update, update.message.text)
    update.message.reply_text('نام محصولی  میخواهید ارسال شود را وارد نمائید: ')
    return B


def get_name_product(update, context):
    
    name_product = update.message.text
    keyboard = [[InlineKeyboardButton("1\u2B50", callback_data='1'),
                 InlineKeyboardButton("2\u2B50 ", callback_data='2')],
                [InlineKeyboardButton("3\u2B50", callback_data='3'),
                 InlineKeyboardButton("4\u2B50 ", callback_data='4'),
                 InlineKeyboardButton("5\u2B50", callback_data='5'),
                 ]]
    global sheet
    max_row = sheet.max_row
    book = load_workbook('dataset.xlsx')
    global bolck
    block = []
    sheets = book.sheetnames
    for i in range(0, len(sheets)):
        if sheets[i] == FileName:
            sheet = book.get_sheet_by_name(FileName)

    for i in range(1, max_row+1):
        if sheet.cell(row=i, column=3).value == name_product:
            block.append(sheet.cell(row=i, column=1).value)
            block.append(sheet.cell(row=i, column=2).value)
            block.append(sheet.cell(row=i, column=3).value)
            block.append(sheet.cell(row=i, column=4).value)
            block.append(sheet.cell(row=i, column=5).value)
            block.append(sheet.cell(row=i, column=6).value)
            block.append(sheet.cell(row=i, column=7).value)
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.bot.send_photo(chat_id=update.message.chat_id, photo=open(r'./image/{}'.format(block[3]), 'rb')
                           ,  caption="کد محصول" + "\t" + '{}'.format(block[0]) + "\n"
                           "نام محصول:   " + "\t" + '{}'.format(block[2]) + "\n"
                           "دسته بندی محصول:  " + "\t" + '{}'.format(block[1]) + "\n"
                        "توضیحات محصول:  "+ "\t" + '{}'.format(block[6]) + "\n"
                         "قیمت محصول:  " + "\t" + '{}'.format(block[4]) + "\n"
                        "لینک پرداخت:   " + "\t" + '{}'.format(block[5]) + "\n"
                          , reply_markup=reply_markup )
   

    return ConversationHandler.END



def cancel(update, context):
    update.message.reply_text('لغو شد', )
    return ConversationHandler.END

if __name__ == '__main__':

    dp = updater.dispatcher
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', id_channel)],

        states={
            A: [MessageHandler(Filters.text & ~Filters.command, get_id_channel)],

            B: [MessageHandler(Filters.text & ~Filters.command, get_name_product)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]

    )
    dp.add_handler(conv_handler)
    updater.start_polling()
    updater.idle()
