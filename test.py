import logging
from telegram import ReplyKeyboardRemove
from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters,
                          ConversationHandler)
from random import randint, seed
from openpyxl import load_workbook
# Enable logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)

IDChannel, NAMEPORDUCT, PHOTO, PRICE, DESCRIPTION, PAY = range(6)
FileName = ""


def make_dataset(filename):
    header = [
        'ProductId', 'Category', 'NameProduct', 'NameImage', 'Price',
        'Link pay', 'Description', 'timestamp'
    ]
    global FileName
    FileName = filename
    global sheet
    book = load_workbook('dataset.xlsx')
    sheets = book.sheetnames
    i = 0
    for i in range(0, len(sheets)):
        if sheets[i] == filename:
            sheet = book.get_sheet_by_name(filename)
            break
    if sheets[i] != filename:
        sheet = book.create_sheet(filename)
    for i in range(1, len(header)):
        sheet.cell(row=1, column=i).value = header[i - 1]
    book.save('dataset.xlsx')


def add_data(header, data):
    book = load_workbook('dataset.xlsx')
    sheet = book.get_sheet_by_name(FileName)
    rowsize = sheet.max_row

    if header == "NameProduct":
        rowsize+=1
        sheet.cell(row=rowsize, column=3).value = data
    if header == "NameImage":
        sheet.cell(row=rowsize, column=4).value = data
    if header == "Price":
        sheet.cell(row=rowsize, column=5).value = data
    if header == "Link pay":
        sheet.cell(row=rowsize, column=6).value = data
    if header == "Description":
        sheet.cell(row=rowsize, column=7).value = data

    book.save('dataset.xlsx')


def id_channel(update, context):
    chat_id = update.effective_chat.id
    context.bot.send_message(chat_id, text="آیدی کانال یا گروه یا شخصی که میخواهید محصولات را برای آن ایجاد کنید بنویسید.")
    return IDChannel


def get_id_channel(update, context):
    logger.info("id_channel of : %s", update.message.text)
    make_dataset(update.message.text)
    update.message.reply_text('نام محصول خود را وارد نمائید:')
    return NAMEPORDUCT


def get_name_product(update, context):
    logger.info("NameProduct: %s", update.message.text)
    add_data("NameProduct", update.message.text)
    update.message.reply_text('عکس محصول خود را وارد نمائید')
    return PHOTO


def get_photo(update, context):
    seed(1)
    val = randint(0, 1000000)
    photo_file = update.message.photo[-1].get_file()
    photo_file.download('{}.jpg'.format(val))
    add_data("NameImage", '{}.jpg'.format(val))
    logger.info("Photo of : %s", val)
    update.message.reply_text('حالا قیمت محصول را وارد نمائید:')
    return PRICE


def get_price(update, context):
    logger.info("Price : %s", update.message.text)
    update.message.reply_text('توضیحات محصول را بنویسید')
    add_data("Price", update.message.text)
    return DESCRIPTION


def get_description(update, context):
    logger.info("Description : %s", update.message.text)
    update.message.reply_text('لینک خرید محصول را وارد نمایید:')
    add_data("Description", update.message.text)
    return PAY


def get_pay(update, context):
    logger.info("Link: %s", update.message.text)
    update.message.reply_text('محصول شما ثبت گردید.')
    add_data("Link pay", update.message.text)
    return ConversationHandler.END


def skip_photo(update, context):
    user = update.message.from_user
    logger.info("User %s did not send a photo.", user.first_name)
    update.message.reply_text('I bet you look great! Now, send me your location please, '
                              'or send /skip.')

    return ConversationHandler.END


def cancel(update, context):
    user = update.message.from_user
    logger.info("User %s canceled the conversation.", user.first_name)
    update.message.reply_text('Bye! I hope we can talk again some day.',
                              reply_markup=ReplyKeyboardRemove())

    return ConversationHandler.END


if __name__ == '__main__':
    updater = Updater('1080958406:AAEjfkoV019XFaDgwnA-9ri89FnBfGphwrE', use_context=True)
    dp = updater.dispatcher
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', id_channel)],

        states={
            IDChannel: [MessageHandler(Filters.text & ~Filters.command, get_id_channel)],

            NAMEPORDUCT: [MessageHandler(Filters.text & ~Filters.command, get_name_product)],

            PHOTO: [MessageHandler(Filters.photo, get_photo),
                    CommandHandler('skip', skip_photo)],

            PRICE: [MessageHandler(Filters.text & ~Filters.command, get_price)],

            DESCRIPTION: [MessageHandler(Filters.text & ~Filters.command, get_description)],

            PAY: [MessageHandler(Filters.text & ~Filters.command, get_pay)],
        },

        fallbacks=[CommandHandler('cancel', cancel)]
    )

    dp.add_handler(conv_handler)
    updater.start_polling()
    updater.idle()


